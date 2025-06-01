import sqlite3
import io
from datetime import datetime, timedelta, date 
from flask import (
    Flask, render_template, request, jsonify, send_file, g, current_app,
    redirect, url_for, flash, session 
)
from werkzeug.security import generate_password_hash, check_password_hash 
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user, login_required, current_user
) 

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os # Adicionado para o caminho do banco de dados (se não estava antes)

# --- Configuração da Aplicação Flask ---
app = Flask(__name__)

# Caminho absoluto para o banco de dados
project_root = os.path.dirname(os.path.abspath(__file__))
app.config['DATABASE'] = os.path.join(project_root, 'sangria_doadores.db')
app.config['SECRET_KEY'] = 'uma_chave_secreta_muito_segura_e_dificil_de_adivinhar_123!@#' # Mantenha uma chave forte!

# --- Configuração do Flask-Login ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' 
login_manager.login_message = "Por favor, faça login para acessar esta página."
login_manager.login_message_category = "error"


# --- Modelo de Usuário (UserMixin para Flask-Login) ---
class User(UserMixin):
    def __init__(self, id, username, full_name, password_hash=None, is_admin=False): 
        self.id = id
        self.username = username
        self.full_name = full_name 
        self.password_hash = password_hash 
        self.is_admin = is_admin

    def get_id(self): 
        return str(self.id)

# Callback para carregar usuário da sessão
@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    user_data = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if user_data:
        return User(
            id=user_data['id'], 
            username=user_data['username'], 
            full_name=user_data['full_name'], 
            password_hash=user_data['password_hash'],
            is_admin=bool(user_data['is_admin']) 
        )
    return None

# --- Funções do Banco de Dados ---
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(
            app.config['DATABASE'], # Já usa o caminho absoluto configurado acima
            detect_types=sqlite3.PARSE_DECLTYPES
        )
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    with current_app.open_resource('schema.sql') as f:
        db.executescript(f.read().decode('utf8'))
    print("Banco de dados inicializado com schema (incluindo tabela users e coluna is_admin).")

@app.cli.command('init-db')
def init_db_command():
    """Limpa os dados existentes e cria novas tabelas."""
    init_db()


# --- Rotas de Autenticação ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index')) 

    if request.method == 'POST':
        full_name = request.form.get('full_name')
        cpf = request.form.get('cpf')
        birth_date_user = request.form.get('birth_date_user')
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if not all([full_name, cpf, birth_date_user, username, password, confirm_password]):
            flash('Todos os campos são obrigatórios!', 'error')
            return redirect(url_for('register'))

        if password != confirm_password:
            flash('As senhas não coincidem!', 'error')
            return redirect(url_for('register'))
        
        if len(password) < 6: 
            flash('A senha deve ter no mínimo 6 caracteres.', 'error')
            return redirect(url_for('register'))

        db = get_db()
        if db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone():
            flash('Este nome de usuário já existe. Escolha outro.', 'error')
            return redirect(url_for('register'))
        if db.execute('SELECT id FROM users WHERE cpf = ?', (cpf,)).fetchone():
            flash('Este CPF já está cadastrado.', 'error')
            return redirect(url_for('register'))

        password_h = generate_password_hash(password)
        
        try:
            db.execute(
                'INSERT INTO users (full_name, cpf, birth_date_user, username, password_hash) VALUES (?, ?, ?, ?, ?)',
                (full_name, cpf, birth_date_user, username, password_h)
            )
            db.commit()
            flash('Cadastro realizado com sucesso! Faça o login.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError: 
            flash('Erro ao cadastrar. Nome de usuário ou CPF podem já estar em uso.', 'error')
            db.rollback()
        except Exception as e:
            current_app.logger.error(f"Erro inesperado no cadastro: {e}")
            flash(f'Ocorreu um erro inesperado ao tentar realizar o cadastro.', 'error')
            db.rollback()
            
        return redirect(url_for('register'))
    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            flash('Usuário e senha são obrigatórios!', 'error')
            return redirect(url_for('login'))

        db = get_db()
        user_data = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()

        if user_data and check_password_hash(user_data['password_hash'], password):
            user_obj = User(
                id=user_data['id'], 
                username=user_data['username'], 
                full_name=user_data['full_name'], 
                password_hash=user_data['password_hash'],
                is_admin=bool(user_data['is_admin']) 
            )
            login_user(user_obj) 
            next_page = request.args.get('next')
            if next_page:
                 return redirect(next_page)
            else:
                 return redirect(url_for('index'))
        else:
            flash('Usuário ou senha inválidos.', 'error')
            return redirect(url_for('login'))
    return render_template('login.html')


@app.route('/logout')
@login_required 
def logout():
    logout_user() 
    flash('Você saiu da sua conta.', 'info')
    return redirect(url_for('login'))


# --- Lógica de Triagem ---
def perform_triage(data):
    today = date.today()
    messages = []
    deferral_days = 0
    overall_status = "apto" 

    try:
        birth_date_obj = datetime.strptime(data.get('birthDate'), '%Y-%m-%d').date()
        age = today.year - birth_date_obj.year - ((today.month, today.day) < (birth_date_obj.month, birth_date_obj.day))
        if not (16 <= age <= 69):
            messages.append(f"Idade ({age} anos) fora do permitido (16-69 anos).")
            overall_status = "inapto_permanente" 
        elif age > 60 and not data.get('lastDonationDate'): 
             if age > 60: 
                messages.append("Primeira doação não permitida acima de 60 anos.")
                overall_status = "inapto_permanente"
    except (ValueError, TypeError):
        messages.append("Data de nascimento inválida.")
        return "inapto_permanente", "Data de nascimento inválida.", 0, None

    try:
        weight = float(str(data.get('weight', '0')).replace(',', '.'))
        if weight < 50:
            messages.append("Peso inferior a 50kg.")
            overall_status = "inapto_permanente" 
    except ValueError:
        messages.append("Valor de peso inválido.")
        overall_status = "inapto_permanente"

    if overall_status != "inapto_permanente":
        if data.get('feverFlu') == 'yes':
            messages.append("Febre/gripe nos últimos 7 dias. Aguardar 7 dias após o fim dos sintomas.")
            deferral_days = max(deferral_days, 7) 
            overall_status = "inapto_temporario"
        if data.get('tattooPiercing') == 'yes':
            messages.append("Tatuagem/piercing nos últimos 6 meses. Aguardar 6 meses.")
            deferral_days = max(deferral_days, 180) 
            overall_status = "inapto_temporario"
        if data.get('hepatitis') == 'yes': 
            messages.append("Hepatite após os 11 anos é um impedimento definitivo.")
            overall_status = "inapto_permanente"
        if data.get('stdPositive') == 'yes': 
            messages.append("Teste positivo para HIV, HTLV, Sífilis ou Hepatite B/C é impedimento definitivo.")
            overall_status = "inapto_permanente"
        if data.get('injectedDrugs') == 'yes':
            messages.append("Uso de drogas injetáveis é impedimento definitivo.")
            overall_status = "inapto_permanente"
        if data.get('pregnantBreastfeeding') == 'yes':
            messages.append("Gestantes ou lactantes (bebê < 12 meses) não podem doar.")
            deferral_days = max(deferral_days, 365) 
            overall_status = "inapto_temporario"
        
    last_donation_str = data.get('lastDonationDate')
    min_interval_days = 90 

    if last_donation_str:
        try:
            last_donation_obj = datetime.strptime(last_donation_str, '%Y-%m-%d').date()
            days_since_last_donation = (today - last_donation_obj).days
            if days_since_last_donation < min_interval_days:
                remaining_days = min_interval_days - days_since_last_donation
                messages.append(f"Ainda faltam {remaining_days} dias desde a última doação (intervalo mínimo {min_interval_days} dias).")
                deferral_days = max(deferral_days, remaining_days)
                overall_status = "inapto_temporario"
        except ValueError:
            messages.append("Data da última doação inválida.")

    final_message = "Triagem concluída. " + (" | ".join(messages) if messages else "Nenhum impedimento direto encontrado nas questões básicas.")
    if overall_status == "apto" and not messages:
        final_message = "Preliminarmente Apto para doação!"

    calculated_next_date = None
    if overall_status == "apto":
        calculated_next_date = today + timedelta(days=min_interval_days) 
        if last_donation_str:
            try:
                last_donation_obj = datetime.strptime(last_donation_str, '%Y-%m-%d').date()
                calculated_next_date = max(calculated_next_date, last_donation_obj + timedelta(days=min_interval_days))
            except ValueError: pass
    elif overall_status == "inapto_temporario":
        if deferral_days > 0 : 
            calculated_next_date = today + timedelta(days=deferral_days)
    return overall_status, final_message, deferral_days, calculated_next_date.isoformat() if calculated_next_date else None


# --- Rotas Principais da Aplicação (Doadores) ---
@app.route('/')
@login_required 
def index():
    return render_template('index.html') 

@app.route('/api/donors', methods=['POST'])
@login_required 
def add_donor():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Nenhum dado enviado"}), 400

    required_fields = ['donorName', 'birthDate', 'weight', 'bloodType']
    for field in required_fields:
        if field not in data or not data[field]:
            return jsonify({"error": f"Campo obrigatório ausente: {field}"}), 400

    triage_status, triage_msg, deferral_d, next_donation_calc = perform_triage(data)

    db = get_db()
    try:
        cursor = db.cursor()
        cursor.execute(
            """
            INSERT INTO donors (
                name, birth_date, weight, blood_type, last_donation_date, contact_info,
                triage_result_status, triage_deferral_days, triage_message, calculated_next_donation_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data.get('donorName'), data.get('birthDate'), float(str(data.get('weight','0')).replace(',','.')),
                data.get('bloodType'), data.get('lastDonationDate') or None, data.get('contactInfo') or None,
                triage_status, deferral_d if triage_status == 'inapto_temporario' and deferral_d > 0 else None,
                triage_msg, next_donation_calc
            )
        )
        db.commit()
        new_donor_id = cursor.lastrowid
    except sqlite3.Error as e:
        db.rollback()
        current_app.logger.error(f"Erro de banco de dados: {e}")
        return jsonify({"error": "Erro ao salvar doador no banco de dados", "details": str(e)}), 500

    return jsonify({
        "message": "Doador processado com sucesso!",
        "id": new_donor_id,
        "triage_status": triage_status,
        "triage_message": triage_msg,
        "deferral_days": deferral_d,
        "calculated_next_donation_date": next_donation_calc
    }), 201


@app.route('/api/donors', methods=['GET'])
@login_required 
def get_donors():
    db = get_db()
    cursor = db.execute('SELECT * FROM donors ORDER BY registration_date DESC')
    donors_list = [dict(row) for row in cursor.fetchall()]

    today_date = date.today()
    for donor in donors_list:
        display_status = donor['triage_result_status']
        if donor['calculated_next_donation_date']:
            try:
                next_date_str = donor['calculated_next_donation_date']
                next_date_obj = None
                if isinstance(next_date_str, str):
                    next_date_obj = date.fromisoformat(next_date_str)
                elif isinstance(next_date_str, date): 
                    next_date_obj = next_date_str
                
                if next_date_obj:
                    if donor['triage_result_status'] == 'inapto_temporario' and next_date_obj <= today_date:
                        display_status = 'apto_pos_espera'
                    elif donor['triage_result_status'] == 'apto' and next_date_obj > today_date:
                        display_status = 'aguardando_intervalo'
            except (ValueError, TypeError) as e:
                current_app.logger.error(f"Erro ao processar data em get_donors: {e} para valor {donor['calculated_next_donation_date']}")
                pass 
        donor['display_status'] = display_status
    
    return jsonify(donors_list)


@app.route('/api/donors/spreadsheet')
@login_required 
def generate_spreadsheet():
    if not current_user.is_admin:
        flash('Acesso negado. Esta funcionalidade é restrita a administradores.', 'error')
        return redirect(url_for('index'))

    db = get_db()
    cursor = db.execute('SELECT * FROM donors ORDER BY name ASC')
    donors_data = [dict(row) for row in cursor.fetchall()]

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Controle de Doações Sangria"

    # --- INÍCIO DAS MODIFICAÇÕES PARA ADICIONAR contact_info ---
    headers = [
        "ID", "Nome", "Data Nasc.", "Peso (kg)", "Tipo Sanguíneo", 
        "Email/Telefone", # << NOVO CABEÇALHO ADICIONADO
        "Última Doação", "Status Triagem", "Dias Inapto", "Mensagem Triagem", 
        "Próxima Data Possível", "Data Cadastro"
    ]
    # --- FIM DAS MODIFICAÇÕES PARA ADICIONAR contact_info ---
    sheet.append(headers)

    fill_apto = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_apto = Font(color="006100", bold=True)
    fill_espera = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    font_espera = Font(color="9C5700", bold=True)
    fill_inapto_temp = PatternFill(start_color="FFD1D1", end_color="FFD1D1", fill_type="solid") 
    font_inapto_temp = Font(color="9C0006", bold=True)
    fill_inapto_perm = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_inapto_perm = Font(color="9C0006", bold=True, italic=True)
    
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = alignment_center
        column_letter = get_column_letter(col_num)
        if header_title == "Mensagem Triagem":
            sheet.column_dimensions[column_letter].width = 50
        # MODIFICADO: Adicionado "Email/Telefone" à lista de colunas com largura 25
        elif header_title in ["Nome", "Próxima Data Possível", "Data Cadastro", "Email/Telefone"]: 
            sheet.column_dimensions[column_letter].width = 25
        else:
            sheet.column_dimensions[column_letter].width = 15

    for donor in donors_data:
        birth_d = donor['birth_date']
        last_d = donor['last_donation_date'] if donor['last_donation_date'] else '-'
        next_d = donor['calculated_next_donation_date'] if donor['calculated_next_donation_date'] else '-'
        
        # ADICIONADO: Pega o dado de contato
        contact = donor['contact_info'] if donor['contact_info'] else '-' 

        registration_value = donor['registration_date'] 
        if isinstance(registration_value, datetime):  
            reg_d = registration_value.strftime('%d/%m/%Y %H:%M')
        elif isinstance(registration_value, str):  
            try:
                reg_d = datetime.fromisoformat(registration_value).strftime('%d/%m/%Y %H:%M')
            except ValueError: 
                reg_d = registration_value 
        else:
            reg_d = '-' 

        # MODIFICADO: Adicionado 'contact' à lista de dados da linha
        row_data = [
            donor['id'], donor['name'], birth_d, donor['weight'], donor['blood_type'],
            contact, # << DADO DE CONTATO ADICIONADO
            last_d, donor['triage_result_status'], 
            donor['triage_deferral_days'] if donor['triage_deferral_days'] else '-',
            donor['triage_message'], next_d, reg_d
        ]
        sheet.append(row_data)
        
        # Lógica de status_cell_idx e alinhamento precisa ser robusta à nova coluna
        try:
            status_cell_idx = headers.index("Status Triagem") + 1
            # ADICIONADO: Índice para a nova coluna de contato para alinhamento (ou certifique-se que a lógica de alinhamento abaixo lida com isso)
            contact_cell_idx = headers.index("Email/Telefone") + 1 
        except ValueError: 
            status_cell_idx = -1 
            contact_cell_idx = -1 # Se o header não for encontrado, não tentaremos alinhar
            
        status_val = (donor['triage_result_status'] or "").lower() 
        
        current_row = sheet.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=current_row, column=col_idx)
            # Lógica de alinhamento MODIFICADA para incluir a nova coluna de contato
            # As colunas que queremos centralizadas: ID, Peso (kg), Dias Inapto
            # As outras (Nome, Data Nasc., Tipo Sanguíneo, Email/Telefone, Última Doação, Status, Mensagem, Próxima Data, Data Cadastro) à esquerda
            if headers[col_idx-1] in ["ID", "Peso (kg)", "Dias Inapto"]:
                 cell.alignment = alignment_center
            else:
                 cell.alignment = alignment_left

            if status_cell_idx > 0 and col_idx == status_cell_idx: 
                if "apto" in status_val:
                    cell.fill = fill_apto
                    cell.font = font_apto
                elif "espera" in status_val or "aguardando" in status_val: 
                    cell.fill = fill_espera
                    cell.font = font_espera
                elif "inapto_temporario" in status_val:
                    cell.fill = fill_inapto_temp
                    cell.font = font_inapto_temp
                elif "inapto_permanente" in status_val:
                    cell.fill = fill_inapto_perm
                    cell.font = font_inapto_perm

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"sangria_controle_doacoes_{date.today().isoformat()}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)